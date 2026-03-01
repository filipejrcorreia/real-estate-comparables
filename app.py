"""
app.py – Real Estate Comparables Search & Reporting Tool
=========================================================
Run with:
    streamlit run app.py
"""

import os
import sqlite3
from datetime import date, datetime

import pandas as pd
import streamlit as st

from comparables_utils import (
    fmt_currency,
    summary_stats,
    to_csv,
    to_excel,
    run_query,
    validate_record,
)
from db_sync import fetch_db, push_db
from build_db import (
    normalize_parish, normalize_type,
    parse_lot_size, clean_yn, to_real, to_int,
)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE  = os.path.join(BASE_DIR, "comparables.db")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Comparables Database",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Ensure DB exists ──────────────────────────────────────────────────────────
def ensure_db():
    """Create an empty database with the correct schema if it doesn't exist."""
    if os.path.exists(DB_FILE):
        return
    conn = sqlite3.connect(DB_FILE)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS comparables (
            id                       INTEGER PRIMARY KEY AUTOINCREMENT,
            property_name            TEXT,
            address                  TEXT,
            parish                   TEXT,
            parish_normalized        TEXT,
            price_sold               REAL,
            sale_date                TEXT,
            sale_year                INTEGER,
            sale_month               INTEGER,
            sq_ft                    REAL,
            beds                     INTEGER,
            baths                    REAL,
            lot_size_raw             TEXT,
            lot_size_acres           REAL,
            no_units                 INTEGER,
            arv                      REAL,
            assessment               REAL,
            property_type            TEXT,
            property_type_normalized TEXT,
            guest                    TEXT,
            pool                     TEXT,
            waterfront               TEXT,
            listed                   TEXT,
            zone                     TEXT,
            notes                    TEXT,
            country                  TEXT DEFAULT 'Bermuda'
        );
        CREATE INDEX IF NOT EXISTS idx_parish ON comparables(parish_normalized);
        CREATE INDEX IF NOT EXISTS idx_type   ON comparables(property_type_normalized);
        CREATE INDEX IF NOT EXISTS idx_date   ON comparables(sale_date);
        CREATE INDEX IF NOT EXISTS idx_price  ON comparables(price_sold);
        CREATE INDEX IF NOT EXISTS idx_sq_ft  ON comparables(sq_ft);
        CREATE INDEX IF NOT EXISTS idx_year   ON comparables(sale_year);
    """)
    conn.commit()
    conn.close()

ensure_db()
fetch_db(DB_FILE)

# ── DB helpers ────────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


@st.cache_data(ttl=300)
def load_lookup_data():
    conn = get_conn()
    parishes = pd.read_sql_query(
        "SELECT DISTINCT parish_normalized v FROM comparables "
        "WHERE parish_normalized IS NOT NULL ORDER BY v",
        conn,
    )["v"].tolist()

    types = pd.read_sql_query(
        "SELECT DISTINCT property_type_normalized v FROM comparables "
        "WHERE property_type_normalized IS NOT NULL ORDER BY v",
        conn,
    )["v"].tolist()

    zones = pd.read_sql_query(
        "SELECT DISTINCT zone v FROM comparables "
        "WHERE zone IS NOT NULL AND length(trim(zone)) BETWEEN 3 AND 10 "
        "ORDER BY v",
        conn,
    )["v"].tolist()

    price_row = pd.read_sql_query(
        "SELECT MIN(price_sold) mn, MAX(price_sold) mx FROM comparables "
        "WHERE price_sold IS NOT NULL",
        conn,
    ).iloc[0]

    date_row = pd.read_sql_query(
        "SELECT MIN(sale_date) mn, MAX(sale_date) mx FROM comparables "
        "WHERE sale_date IS NOT NULL",
        conn,
    ).iloc[0]

    sqft_row = pd.read_sql_query(
        "SELECT MIN(CAST(sq_ft AS REAL)) mn, MAX(CAST(sq_ft AS REAL)) mx "
        "FROM comparables "
        "WHERE sq_ft IS NOT NULL AND typeof(sq_ft) IN ('real','integer')",
        conn,
    ).iloc[0]

    lot_row = pd.read_sql_query(
        "SELECT MIN(CAST(lot_size_acres AS REAL)) mn, MAX(CAST(lot_size_acres AS REAL)) mx "
        "FROM comparables "
        "WHERE lot_size_acres IS NOT NULL AND typeof(lot_size_acres) IN ('real','integer')",
        conn,
    ).iloc[0]

    conn.close()
    return parishes, types, zones, price_row, date_row, sqft_row, lot_row


def query(filters: dict) -> pd.DataFrame:
    """Run a filtered query against the database and return a DataFrame."""
    conn = get_conn()
    df = run_query(conn, filters)
    conn.close()
    return df


# fmt_currency, summary_stats, to_csv, to_excel imported from comparables_utils


# ═══════════════════════════════════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════════════════════════════════

parishes, types, zones, price_row, date_row, sqft_row, lot_row = load_lookup_data()

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/color/96/real-estate.png", width=56)
    st.title("Search Filters")
    st.caption("Leave a field blank to include all values.")

    # Text search
    search_text = st.text_input("🔍 Property name / address", "")

    st.divider()
    st.subheader("📍 Location")

    selected_parishes = st.multiselect("Parish / Region", parishes)
    country_opt = st.selectbox("Country", ["All", "Bermuda"])

    st.divider()
    st.subheader("🏠 Property")

    selected_types = st.multiselect("Property Type", types)
    selected_zones = st.multiselect("Zone Code", zones)

    col_g, col_p = st.columns(2)
    guest_opt      = col_g.selectbox("Guest Apt", ["All", "Y", "N"])
    pool_opt       = col_p.selectbox("Pool", ["All", "Y", "N"])
    col_w, col_l   = st.columns(2)
    waterfront_opt = col_w.selectbox("Waterfront", ["All", "Y", "N"])
    listed_opt     = col_l.selectbox("Listed", ["All", "Y", "N"])

    st.divider()
    st.subheader("💰 Price (USD)")

    try:
        price_min_v = int(float(price_row["mn"])) if price_row["mn"] not in (None, "") else 0
    except (ValueError, TypeError):
        price_min_v = 0
    try:
        price_max_v = int(float(price_row["mx"])) if price_row["mx"] not in (None, "") else 10_000_000
    except (ValueError, TypeError):
        price_max_v = 10_000_000
    price_range = st.slider(
        "Sale Price Range",
        min_value=0,
        max_value=price_max_v,
        value=(price_min_v, price_max_v),
        step=10_000,
        format="$%d",
    )

    st.divider()
    st.subheader("📅 Sale Date")

    min_date = date(2020, 1, 1)
    max_date = date.today()
    if date_row["mn"]:
        try:
            min_date = pd.to_datetime(str(date_row["mn"])).date()
        except Exception:
            pass
    if date_row["mx"]:
        try:
            max_date = pd.to_datetime(str(date_row["mx"])).date()
        except Exception:
            pass

    date_from = st.date_input("From", value=min_date,
                              min_value=min_date, max_value=max_date)
    date_to   = st.date_input("To",   value=max_date,
                              min_value=min_date, max_value=max_date)

    st.divider()
    st.subheader("📐 Size / Lot")

    try:
        sqft_max_v = int(float(sqft_row["mx"])) if sqft_row["mx"] not in (None, "") else 20_000
    except (ValueError, TypeError):
        sqft_max_v = 20_000
    sqft_range = st.slider("Sq. Ft. Range", 0, sqft_max_v,
                            (0, sqft_max_v), step=100)

    try:
        lot_max_v = round(float(lot_row["mx"]), 2) if lot_row["mx"] not in (None, "") else 10.0
    except (ValueError, TypeError):
        lot_max_v = 10.0
    lot_range  = st.slider("Lot Size (acres)", 0.0, lot_max_v,
                           (0.0, lot_max_v), step=0.01)

    st.divider()
    st.subheader("🛏 Beds / Baths / Units")

    c1, c2 = st.columns(2)
    beds_min  = c1.number_input("Beds min",  min_value=0, value=0, step=1)
    beds_max  = c2.number_input("Beds max",  min_value=0, value=10, step=1)
    baths_min = c1.number_input("Baths min", min_value=0, value=0, step=1)
    c_u1, c_u2 = st.columns(2)
    units_min = c_u1.number_input("Units min", min_value=0, value=0, step=1)
    units_max = c_u2.number_input("Units max", min_value=0, value=50, step=1)

    st.divider()
    run_search = st.button("🔎 Search", width="stretch", type="primary")
    clear      = st.button("↺ Reset Filters", width="stretch")

# ── Reset ─────────────────────────────────────────────────────────────────────
if clear:
    st.rerun()

# ── Build filters dict ────────────────────────────────────────────────────────
filters: dict = {
    "search":     search_text.strip() or None,
    "parishes":   selected_parishes or None,
    "types":      selected_types    or None,
    "zones":      selected_zones    or None,
    "country":    None if country_opt == "All" else country_opt,
    "price_min":  price_range[0] if price_range[0] > 0          else None,
    "price_max":  price_range[1] if price_range[1] < price_max_v else None,
    "date_from":  date_from,
    "date_to":    date_to,
    "sqft_min":   sqft_range[0] if sqft_range[0] > 0            else None,
    "sqft_max":   sqft_range[1] if sqft_range[1] < sqft_max_v   else None,
    "lot_min":    lot_range[0]  if lot_range[0]  > 0            else None,
    "lot_max":    lot_range[1]  if lot_range[1]  < lot_max_v    else None,
    "beds_min":   beds_min  if beds_min  > 0  else None,
    "beds_max":   beds_max  if beds_max  < 10 else None,
    "baths_min":  baths_min if baths_min > 0  else None,
    "units_min":  units_min if units_min > 0  else None,
    "units_max":  units_max if units_max < 50 else None,
    "pool":       pool_opt,
    "guest":      guest_opt,
    "waterfront": waterfront_opt,
    "listed":     listed_opt,
}

# ── Always show results (auto-search) ─────────────────────────────────────────
df = query(filters)
stats = summary_stats(df)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("## 🏠 Real Estate Comparables Database")
st.caption(
    "Comparable sales database · "
    f"Database: **{DB_FILE.split('/')[-1]}** · "
    "Use the **Add Record** tab to enter new comparables."
)

# ── KPI strip ─────────────────────────────────────────────────────────────────
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Matching Records",   f"{stats['count']:,}")
k2.metric("Avg Sale Price",
          fmt_currency(stats["avg_price"]) if stats["avg_price"] else "—")
k3.metric("Median Sale Price",
          fmt_currency(stats["med_price"]) if stats["med_price"] else "—")
k4.metric("Min Price",
          fmt_currency(stats["min_price"]) if stats["min_price"] else "—")
k5.metric("Max Price",
          fmt_currency(stats["max_price"]) if stats["max_price"] else "—")

st.divider()

# ── Tabs: Results | Charts | Export ──────────────────────────────────────────
tab_res, tab_chart, tab_export, tab_add, tab_import = st.tabs(
    ["📋 Results Table", "📊 Charts & Analytics", "⬇️ Export / Report", "➕ Add Record", "📥 Import File"]
)

# ────────────────────────────────────────────────────────── Results Table ─────
with tab_res:
    if df.empty:
        st.info("No records match the current filters.")
    else:
        # Display columns (drop raw IDs / technical cols)
        display_cols = [
            "property_name", "address", "parish", "type",
            "price_sold", "sale_date",
            "sq_ft", "beds", "baths",
            "lot_size", "lot_size_acres", "units",
            "arv", "assessment",
            "guest", "pool", "waterfront", "listed",
            "zone", "notes", "country",
        ]
        display_cols = [c for c in display_cols if c in df.columns]
        show_df = df[display_cols].copy()

        # Friendly column names
        rename = {
            "property_name": "Property",
            "address":       "Address",
            "parish":        "Parish",
            "type":          "Type",
            "price_sold":    "Sale Price ($)",
            "sale_date":     "Date",
            "sq_ft":         "Sq. Ft.",
            "beds":          "Beds",
            "baths":         "Baths",
            "lot_size":      "Lot Size",
            "lot_size_acres":"Lot (ac)",
            "units":         "Units",
            "arv":           "ARV ($)",
            "assessment":    "Assessment",
            "guest":         "Guest",
            "pool":          "Pool",
            "waterfront":    "Waterfront",
            "listed":        "Listed",
            "zone":          "Zone",
            "notes":         "Notes",
            "country":       "Country",
        }
        show_df.rename(columns=rename, inplace=True)

        st.dataframe(
            show_df,
            width="stretch",
            height=520,
            column_config={
                "Sale Price ($)": st.column_config.NumberColumn(format="$%d"),
                "ARV ($)":        st.column_config.NumberColumn(format="$%d"),
                "Assessment":     st.column_config.NumberColumn(format="$%d"),
                "Lot (ac)":       st.column_config.NumberColumn(format="%.3f"),
            },
        )
        st.caption(f"Showing {len(df):,} record(s)")

# ────────────────────────────────────────────────────────── Charts ────────────
with tab_chart:
    if df.empty:
        st.info("No data to chart.")
    else:
        # Try to import plotly; fall back to st native charts
        try:
            import plotly.express as px

            c1, c2 = st.columns(2)

            # Price distribution
            price_df = df["price_sold"].dropna()
            if not price_df.empty:
                with c1:
                    st.subheader("Sale Price Distribution")
                    fig = px.histogram(price_df, nbins=30, labels={"value": "Price ($)"},
                                       color_discrete_sequence=["#1F3864"])
                    fig.update_layout(showlegend=False, margin=dict(t=20))
                    st.plotly_chart(fig, width="stretch")

            # Count by Parish
            parish_counts = df["parish"].value_counts().reset_index()
            parish_counts.columns = ["Parish", "Count"]
            with c2:
                st.subheader("Records by Parish")
                fig2 = px.bar(parish_counts, x="Parish", y="Count",
                              color="Parish",
                              color_discrete_sequence=px.colors.qualitative.Set2)
                fig2.update_layout(showlegend=False, margin=dict(t=20),
                                   xaxis_tickangle=-40)
                st.plotly_chart(fig2, width="stretch")

            c3, c4 = st.columns(2)

            # Count by Type
            type_counts = df["type"].value_counts().reset_index()
            type_counts.columns = ["Type", "Count"]
            with c3:
                st.subheader("Records by Property Type")
                fig3 = px.pie(type_counts, names="Type", values="Count",
                              hole=0.4)
                fig3.update_layout(margin=dict(t=20))
                st.plotly_chart(fig3, width="stretch")

            # Average price over time (by year)
            yr_df = df.dropna(subset=["price_sold", "sale_date"]).copy()
            yr_df["Year"] = pd.to_datetime(yr_df["sale_date"]).dt.year
            avg_yr = yr_df.groupby("Year")["price_sold"].median().reset_index()
            avg_yr.columns = ["Year", "Median Price"]
            if len(avg_yr) > 1:
                with c4:
                    st.subheader("Median Sale Price by Year")
                    fig4 = px.line(avg_yr, x="Year", y="Median Price",
                                   markers=True,
                                   color_discrete_sequence=["#1F3864"])
                    fig4.update_layout(margin=dict(t=20))
                    st.plotly_chart(fig4, width="stretch")

            # Price vs Sq Ft scatter
            scatter_df = df.dropna(subset=["price_sold", "sq_ft"])
            if not scatter_df.empty:
                st.subheader("Sale Price vs. Sq. Ft.")
                fig5 = px.scatter(
                    scatter_df, x="sq_ft", y="price_sold",
                    color="type" if "type" in scatter_df.columns else None,
                    hover_data=["property_name", "parish", "sale_date"],
                    labels={"sq_ft": "Sq. Ft.", "price_sold": "Sale Price ($)"},
                    opacity=0.7,
                )
                fig5.update_layout(margin=dict(t=20))
                st.plotly_chart(fig5, width="stretch")

        except ImportError:
            # Fallback – native Streamlit charts
            st.warning("Install plotly for richer charts:  pip install plotly")
            by_parish = df["parish"].value_counts()
            st.bar_chart(by_parish)
            price_data = df.set_index("sale_date")["price_sold"].dropna()
            st.line_chart(price_data)

# ────────────────────────────────────────────────────────── Export ────────────
with tab_export:
    st.subheader("Export Results")

    if df.empty:
        st.info("Run a search first.")
    else:
        st.write(f"**{len(df):,} records** ready for export.")

        col_csv, col_xlsx = st.columns(2)

        with col_csv:
            st.download_button(
                label="⬇️ Download CSV",
                data=to_csv(df),
                file_name=f"comparables_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                width="stretch",
            )

        with col_xlsx:
            xlsx_bytes = to_excel(df, stats)
            st.download_button(
                label="⬇️ Download Excel Report",
                data=xlsx_bytes,
                file_name=f"comparables_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

        st.divider()
        st.subheader("📑 Summary Statistics")

        stat_table = {
            "Metric": [
                "Total Records Shown",
                "Records with Sale Price",
                "Minimum Sale Price",
                "Maximum Sale Price",
                "Average Sale Price",
                "Median Sale Price",
                "Average Sq. Ft.",
                "Avg Price per Sq. Ft.",
            ],
            "Value": [
                f"{stats['count']:,}",
                f"{stats['with_price']:,}",
                fmt_currency(stats["min_price"])  if stats["min_price"]  else "—",
                fmt_currency(stats["max_price"])  if stats["max_price"]  else "—",
                fmt_currency(stats["avg_price"])  if stats["avg_price"]  else "—",
                fmt_currency(stats["med_price"])  if stats["med_price"]  else "—",
                f"{stats['avg_sqft']:,.0f}" if stats["avg_sqft"] else "—",
                fmt_currency(stats["price_per_sqft"]) if stats["price_per_sqft"] else "—",
            ],
        }
        st.table(pd.DataFrame(stat_table))

        # Breakdown tables
        r1, r2 = st.columns(2)
        with r1:
            st.markdown("**By Parish**")
            p = df["parish"].value_counts().rename_axis("Parish").reset_index(name="Count")
            st.dataframe(p, width="stretch", hide_index=True)
        with r2:
            st.markdown("**By Property Type**")
            t = df["type"].value_counts().rename_axis("Type").reset_index(name="Count")
            st.dataframe(t, width="stretch", hide_index=True)

# ────────────────────────────────────────────────── Add Record ────
with tab_add:
    st.subheader("➕ Add New Comparable")
    st.caption("Fields marked * are required. For Parish and Type, pick from the list or enter a custom value.")

    with st.form("add_record_form", clear_on_submit=True):

        st.markdown("**📍 Property Details**")
        ca, cb = st.columns(2)
        prop_name = ca.text_input("Property Name *", placeholder="e.g. Sunrise Villa")
        address   = cb.text_input("Address",         placeholder="e.g. 12 Harbour Road")

        cc, cd = st.columns(2)
        parish_sel    = cc.selectbox("Parish",                    [""] + parishes + ["➕ Other…"])
        parish_custom = cc.text_input("Custom parish (if Other)",  placeholder="Type here")
        type_sel      = cd.selectbox("Property Type",             [""] + types    + ["➕ Other…"])
        type_custom   = cd.text_input("Custom type (if Other)",    placeholder="Type here")

        ce, cf = st.columns(2)
        country = ce.text_input("Country",   value="Bermuda")
        zone    = cf.text_input("Zone Code", placeholder="e.g. 110")

        st.divider()
        st.markdown("**💰 Financials**")
        cg, ch = st.columns(2)
        price_sold  = cg.number_input("Sale Price ($)",  min_value=0, value=None, step=1000, format="%d")
        sale_date   = ch.date_input(  "Sale Date",       value=None)
        arv         = cg.number_input("ARV ($)",         min_value=0, value=None, step=1000, format="%d")
        assessment  = ch.number_input("Assessment ($)",  min_value=0, value=None, step=1000, format="%d")

        st.divider()
        st.markdown("**📐 Size & Features**")
        ci, cj, ck = st.columns(3)
        sq_ft          = ci.number_input("Sq. Ft.",           min_value=0,   value=None, step=10,   format="%d")
        beds           = cj.number_input("Beds",              min_value=0,   value=None, step=1,    format="%d")
        baths          = ck.number_input("Baths",             min_value=0.0, value=None, step=0.5)
        units          = ci.number_input("No. Units",         min_value=0,   value=None, step=1,    format="%d")
        lot_size_text  = cj.text_input(  "Lot Size (text)",   placeholder="e.g. 0.25 Acres")
        lot_size_acres = ck.number_input("Lot Size (acres)",  min_value=0.0, value=None, step=0.01, format="%.3f")

        st.divider()
        st.markdown("**✅ Flags**")
        cl, cm, cn, co2 = st.columns(4)
        guest_v      = cl.selectbox("Guest Apt",  ["", "Y", "N"])
        pool_v       = cm.selectbox("Pool",       ["", "Y", "N"])
        waterfront_v = cn.selectbox("Waterfront", ["", "Y", "N"])
        listed_v     = co2.selectbox("Listed",    ["", "Y", "N"])

        notes = st.text_area("Notes", placeholder="Any additional notes…")

        submitted = st.form_submit_button("💾 Save Record", type="primary", width="stretch")

        if submitted:
            vresult = validate_record({
                "property_name":  prop_name,
                "price_sold":     price_sold,
                "sale_date":      sale_date,
                "sq_ft":          sq_ft,
                "beds":           beds,
                "baths":          baths,
                "lot_size_acres": lot_size_acres,
                "units":          units,
                "country":        country,
            })

            for err in vresult["errors"]:
                st.error(f"❌ {err}")
            for wrn in vresult["warnings"]:
                st.warning(f"⚠️ {wrn}")

            if vresult["errors"]:
                st.stop()

            if parish_sel == "➕ Other…" and not parish_custom.strip():
                st.error("❌ Please enter a parish name in the custom field.")
                st.stop()
            if type_sel == "➕ Other…" and not type_custom.strip():
                st.error("❌ Please enter a property type in the custom field.")
                st.stop()
            else:
                parish_val = (parish_custom.strip() if parish_sel == "➕ Other…" else parish_sel) or None
                type_val   = (type_custom.strip()   if type_sel   == "➕ Other…" else type_sel)   or None

                sd_str  = sale_date.isoformat() if sale_date else None
                sd_year = sale_date.year         if sale_date else None
                sd_mon  = sale_date.month        if sale_date else None

                _conn = get_conn()
                try:
                    _conn.execute("""
                        INSERT INTO comparables (
                            property_name, address,
                            parish, parish_normalized,
                            price_sold, sale_date, sale_year, sale_month,
                            sq_ft, beds, baths,
                            lot_size_raw, lot_size_acres,
                            no_units, arv, assessment,
                            property_type, property_type_normalized,
                            guest, pool, waterfront, listed,
                            zone, notes, country
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        prop_name.strip(),
                        address.strip()   or None,
                        parish_val, parish_val,
                        price_sold,
                        sd_str, sd_year, sd_mon,
                        sq_ft, beds, baths,
                        lot_size_text.strip() or None,
                        lot_size_acres,
                        units, arv, assessment,
                        type_val, type_val,
                        guest_v      or None,
                        pool_v       or None,
                        waterfront_v or None,
                        listed_v     or None,
                        zone.strip()  or None,
                        notes.strip() or None,
                        country.strip() or "Bermuda",
                    ))
                    _conn.commit()
                    st.success(f"✅ **{prop_name.strip()}** saved successfully!")
                    push_db(DB_FILE, f"Add record: {prop_name.strip()}")
                    load_lookup_data.clear()
                    st.rerun()
                except Exception as exc:
                    st.error(f"Failed to save: {exc}")
                finally:
                    _conn.close()

# ──────────────────────────────────────────────────────── Import File ─────────
with tab_import:
    st.subheader("📥 Batch Import from Excel")
    st.caption(
        "Upload an `.xlsx` or `.xlsm` file. "
        "The sheet must follow the standard 19-column layout (same as the original spreadsheet). "
        "Duplicate records are **not** automatically detected — review the preview before importing."
    )

    uploaded = st.file_uploader(
        "Choose a spreadsheet",
        type=["xlsx", "xlsm"],
        label_visibility="collapsed",
    )

    if uploaded:
        import openpyxl as _openpyxl
        import io as _io

        try:
            wb_up = _openpyxl.load_workbook(
                _io.BytesIO(uploaded.read()), read_only=True, keep_vba=False
            )
        except Exception as e:
            st.error(f"Could not open file: {e}")
            wb_up = None

        if wb_up:
            sheet_names = wb_up.sheetnames
            chosen_sheet = st.selectbox(
                "Sheet to import",
                sheet_names,
                index=sheet_names.index("Sheet1") if "Sheet1" in sheet_names else 0,
            )

            import_mode = st.radio(
                "Import mode",
                ["Append to existing records", "Replace ALL existing records"],
                captions=[
                    "New rows are added; current records are kept.",
                    "⚠️ Deletes every current record before importing.",
                ],
            )

            ws_up = wb_up[chosen_sheet]
            all_rows = [r for r in ws_up.iter_rows(values_only=True)
                        if any(v is not None for v in r)]

            if len(all_rows) < 2:
                st.warning("Sheet appears to be empty or has no data rows.")
            else:
                header_up = all_rows[0]
                data_up   = all_rows[1:]
                col_up    = {str(name).strip(): idx
                             for idx, name in enumerate(header_up) if name}

                def _g(row, name, default=None):
                    idx = col_up.get(name)
                    if idx is None or idx >= len(row):
                        return default
                    v = row[idx]
                    return v if v is not None else default

                EXPECTED = [
                    "Property Name", "Address", "Parish", "Price Sold",
                    "Date", "Sq. ft.", "Bed", "Bath", "Lot Size",
                    "No. Units", "ARV", "Assessment", "Type",
                    "Guest", "Pool", "Waterfront", "Listed", "Zone", "Notes",
                ]
                missing_cols = [c for c in EXPECTED if c not in col_up]
                if missing_cols:
                    st.warning(
                        f"⚠️ {len(missing_cols)} expected column(s) not found in this sheet: "
                        + ", ".join(f"`{c}`" for c in missing_cols)
                        + "  \nRecognised columns will still be imported; missing ones will be blank."
                    )

                # Build preview DataFrame
                preview_records = []
                for row in data_up[:10]:
                    pname = _g(row, "Property Name")
                    price = to_real(_g(row, "Price Sold"))
                    parish_raw = _g(row, "Parish")
                    type_raw   = _g(row, "Type")
                    preview_records.append({
                        "Property Name":   str(pname) if pname else "",
                        "Parish":          normalize_parish(str(parish_raw)) if parish_raw else "",
                        "Type":            normalize_type(str(type_raw))   if type_raw   else "",
                        "Price Sold ($)": f"${price:,.0f}" if price else "",
                        "Date":            str(_g(row, "Date") or "")[:10],
                        "Sq. ft.":         str(to_real(_g(row, "Sq. ft.")) or ""),
                    })

                st.markdown(f"**{len(data_up):,} data rows detected** — preview of first 10:")
                st.dataframe(pd.DataFrame(preview_records), width="stretch", hide_index=True)

                # ── Validation report ──────────────────────────────────────
                flagged = []
                for i, row in enumerate(data_up):
                    rv = validate_record({
                        "property_name":  _g(row, "Property Name"),
                        "price_sold":     to_real(_g(row, "Price Sold")),
                        "sale_date":      _g(row, "Date"),
                        "sq_ft":          to_real(_g(row, "Sq. ft.")),
                        "beds":           to_int(_g(row, "Bed")),
                        "baths":          to_real(_g(row, "Bath")),
                        "lot_size_acres": to_real(_g(row, "Lot Size")),
                        "units":          to_int(_g(row, "No. Units")),
                    })
                    all_issues = rv["errors"] + rv["warnings"]
                    if all_issues:
                        flagged.append({
                            "Row":           i + 2,   # +2: 1-based + header row
                            "Property Name": str(_g(row, "Property Name") or ""),
                            "Issues":        " | ".join(all_issues),
                            "Severity":      "❌ Error" if rv["errors"] else "⚠️ Warning",
                        })

                rows_with_errors = sum(1 for f in flagged if f["Severity"] == "❌ Error")

                if flagged:
                    with st.expander(
                        f"🔍 Validation report: {len(flagged)} row(s) flagged "
                        f"({rows_with_errors} error(s), {len(flagged) - rows_with_errors} warning(s))",
                        expanded=rows_with_errors > 0,
                    ):
                        st.caption(
                            "Rows with **errors** will be imported with invalid fields set to blank. "
                            "Rows with **warnings** will still be imported as-is."
                        )
                        st.dataframe(
                            pd.DataFrame(flagged),
                            width="stretch",
                            hide_index=True,
                            column_config={
                                "Row":      st.column_config.NumberColumn(width="small"),
                                "Severity": st.column_config.TextColumn(width="small"),
                            },
                        )
                else:
                    st.success("✅ All rows passed validation.")

                confirm = st.button(
                    f"{'⚠️ Replace all &amp; import' if 'Replace' in import_mode else '📥 Import'} "
                    f"{len(data_up):,} records",
                    type="primary",
                    width="stretch",
                )

                if confirm:
                    progress = st.progress(0, text="Importing…")
                    _conn2 = get_conn()
                    try:
                        if "Replace" in import_mode:
                            _conn2.execute("DELETE FROM comparables")

                        inserted = 0
                        total    = len(data_up)

                        for i, row in enumerate(data_up):
                            # Use "(Unnamed)" instead of skipping rows with no property name
                            pname = str(_g(row, "Property Name") or "").strip() or "(Unnamed)"

                            parish_raw = str(_g(row, "Parish") or "")
                            type_raw   = str(_g(row, "Type")   or "")
                            date_raw   = _g(row, "Date")

                            if hasattr(date_raw, "strftime"):
                                sd_str  = date_raw.strftime("%Y-%m-%d")
                                sd_year = date_raw.year
                                sd_mon  = date_raw.month
                            elif date_raw:
                                sd_str  = str(date_raw)[:10]
                                try:
                                    _d = date.fromisoformat(sd_str)
                                    sd_year, sd_mon = _d.year, _d.month
                                except ValueError:
                                    sd_year = sd_mon = None
                            else:
                                sd_str = sd_year = sd_mon = None

                            lot_raw, lot_acres = parse_lot_size(_g(row, "Lot Size"))

                            # Sanitize: replace negative numeric values with NULL
                            # (invalid values become blank rather than excluding the row)
                            price_val = to_real(_g(row, "Price Sold"))
                            sqft_val  = to_real(_g(row, "Sq. ft."))
                            beds_val  = to_int(_g(row, "Bed"))
                            baths_val = to_real(_g(row, "Bath"))
                            units_val = to_int(_g(row, "No. Units"))
                            if price_val  is not None and price_val  < 0: price_val  = None
                            if sqft_val   is not None and sqft_val   < 0: sqft_val   = None
                            if beds_val   is not None and beds_val   < 0: beds_val   = None
                            if baths_val  is not None and baths_val  < 0: baths_val  = None
                            if lot_acres  is not None and lot_acres  < 0: lot_acres  = None
                            if units_val  is not None and units_val  < 0: units_val  = None

                            _conn2.execute("""
                                INSERT INTO comparables (
                                    property_name, address,
                                    parish, parish_normalized,
                                    price_sold, sale_date, sale_year, sale_month,
                                    sq_ft, beds, baths,
                                    lot_size_raw, lot_size_acres,
                                    no_units, arv, assessment,
                                    property_type, property_type_normalized,
                                    guest, pool, waterfront, listed,
                                    zone, notes, country
                                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """, (
                                pname,
                                str(_g(row, "Address") or "").strip() or None,
                                parish_raw or None,
                                normalize_parish(parish_raw) if parish_raw else None,
                                price_val,
                                sd_str, sd_year, sd_mon,
                                sqft_val,
                                beds_val,
                                baths_val,
                                lot_raw, lot_acres,
                                units_val,
                                to_real(_g(row, "ARV")),
                                to_real(_g(row, "Assessment")),
                                type_raw or None,
                                normalize_type(type_raw) if type_raw else None,
                                clean_yn(_g(row, "Guest")),
                                clean_yn(_g(row, "Pool")),
                                clean_yn(_g(row, "Waterfront")),
                                clean_yn(_g(row, "Listed")),
                                str(_g(row, "Zone") or "").strip() or None,
                                str(_g(row, "Notes") or "").strip() or None,
                                "Bermuda",
                            ))
                            inserted += 1
                            if i % 50 == 0:
                                progress.progress(
                                    min(i / total, 0.99),
                                    text=f"Importing… {i:,}/{total:,}"
                                )

                        _conn2.commit()
                        progress.progress(1.0, text="Done!")
                        action = "replaced with" if "Replace" in import_mode else "added —"
                        st.success(
                            f"✅ Import complete: **{inserted:,}** records {action}."
                        )
                        push_db(DB_FILE, f"Batch import: {inserted} records ({import_mode.split()[0].lower()})")
                        load_lookup_data.clear()
                        st.rerun()

                    except Exception as exc:
                        _conn2.rollback()
                        st.error(f"Import failed and was rolled back: {exc}")
                    finally:
                        _conn2.close()
