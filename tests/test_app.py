"""
test_app.py – Unit tests for comparables_utils.py
==================================================
Run:  pytest tests/test_app.py -v
"""

import io
import sqlite3
import statistics
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from comparables_utils import (
    fmt_currency, summary_stats, to_csv, to_excel,
    build_query_sql, run_query,
)
from tests.conftest import SAMPLE_ROWS


# ═══════════════════════════════════════════════════════════════════════════════
# fmt_currency
# ═══════════════════════════════════════════════════════════════════════════════

class TestFmtCurrency:
    def test_integer_value(self):
        assert fmt_currency(1_000_000) == "$1,000,000"

    def test_float_rounds_down(self):
        assert fmt_currency(999_999.49) == "$999,999"

    def test_float_rounds_up(self):
        assert fmt_currency(999_999.50) == "$1,000,000"

    def test_zero(self):
        assert fmt_currency(0) == "$0"

    def test_nan_returns_empty(self):
        assert fmt_currency(float("nan")) == ""

    def test_none_returns_empty(self):
        # pd.isna(None) is True
        assert fmt_currency(None) == ""

    def test_small_value(self):
        assert fmt_currency(500) == "$500"


# ═══════════════════════════════════════════════════════════════════════════════
# summary_stats
# ═══════════════════════════════════════════════════════════════════════════════

class TestSummaryStats:
    def test_normal_dataframe(self, sample_df):
        stats = summary_stats(sample_df)

        # 6 rows total
        assert stats["count"] == 6

        # 5 have a price_sold (one is None)
        assert stats["with_price"] == 5

        prices = [980_000, 650_000, 1_200_000, 420_000, 1_500_000]
        assert stats["min_price"] == pytest.approx(min(prices))
        assert stats["max_price"] == pytest.approx(max(prices))
        assert stats["avg_price"] == pytest.approx(sum(prices) / len(prices))

        assert stats["med_price"] == pytest.approx(statistics.median(prices))

    def test_sqft_average(self, sample_df):
        stats = summary_stats(sample_df)
        # sq_ft values: 2500, 1800, 3200, 900, 1400 (None excluded)
        expected_avg = sum([2500, 1800, 3200, 900, 1400]) / 5
        assert stats["avg_sqft"] == pytest.approx(expected_avg)

    def test_price_per_sqft(self, sample_df):
        stats = summary_stats(sample_df)
        # Only rows with both price_sold AND sq_ft > 0
        # (980000/2500), (650000/1800), (1200000/3200), (420000/900)
        # "Commercial Bld" has sq_ft=None; "No Price Prop" has price=None
        pairs = [(980_000, 2500), (650_000, 1800), (1_200_000, 3200), (420_000, 900)]
        expected = sum(p / s for p, s in pairs) / len(pairs)
        assert stats["price_per_sqft"] == pytest.approx(expected, rel=1e-3)

    def test_empty_dataframe(self):
        empty = pd.DataFrame(columns=["price_sold", "sq_ft"])
        stats = summary_stats(empty)

        assert stats["count"] == 0
        assert stats["with_price"] == 0
        assert stats["min_price"]      is None
        assert stats["max_price"]      is None
        assert stats["avg_price"]      is None
        assert stats["med_price"]      is None
        assert stats["avg_sqft"]       is None
        assert stats["price_per_sqft"] is None

    def test_all_prices_null(self, sample_df):
        df = sample_df.copy()
        df["price_sold"] = None
        stats = summary_stats(df)
        assert stats["with_price"] == 0
        assert stats["min_price"]  is None
        assert stats["avg_price"]  is None

    def test_single_row(self):
        df = pd.DataFrame({"price_sold": [500_000], "sq_ft": [1500]})
        stats = summary_stats(df)
        assert stats["count"]      == 1
        assert stats["min_price"]  == pytest.approx(500_000)
        assert stats["max_price"]  == pytest.approx(500_000)
        assert stats["avg_price"]  == pytest.approx(500_000)
        assert stats["med_price"]  == pytest.approx(500_000)
        assert stats["price_per_sqft"] == pytest.approx(500_000 / 1500)


# ═══════════════════════════════════════════════════════════════════════════════
# to_csv
# ═══════════════════════════════════════════════════════════════════════════════

class TestToCsv:
    def test_returns_bytes(self, sample_df):
        result = to_csv(sample_df)
        assert isinstance(result, bytes)

    def test_utf8_encoded(self, sample_df):
        result = to_csv(sample_df)
        text = result.decode("utf-8")
        assert "property_name" in text

    def test_header_row_present(self, sample_df):
        text = to_csv(sample_df).decode("utf-8")
        first_line = text.splitlines()[0]
        for col in ("property_name", "price_sold", "parish"):
            assert col in first_line

    def test_data_rows_present(self, sample_df):
        text = to_csv(sample_df).decode("utf-8")
        assert "Sunrise Villa" in text
        assert "980000" in text

    def test_row_count(self, sample_df):
        text = to_csv(sample_df).decode("utf-8")
        # header + 6 data rows = 7 lines (last line may be empty)
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) == 7

    def test_empty_dataframe(self):
        df = pd.DataFrame(columns=["property_name", "price_sold"])
        result = to_csv(df)
        text = result.decode("utf-8")
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) == 1   # header only


# ═══════════════════════════════════════════════════════════════════════════════
# to_excel
# ═══════════════════════════════════════════════════════════════════════════════

class TestToExcel:
    def _make_stats(self, df):
        return summary_stats(df)

    def test_returns_bytes(self, sample_df):
        stats = self._make_stats(sample_df)
        result = to_excel(sample_df, stats)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_magic_bytes(self, sample_df):
        """Excel (.xlsx) files start with the ZIP PK header."""
        result = to_excel(sample_df, self._make_stats(sample_df))
        assert result[:2] == b"PK"

    def test_contains_two_sheets(self, sample_df):
        import openpyxl
        result = to_excel(sample_df, self._make_stats(sample_df))
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Comparables" in wb.sheetnames
        assert "Summary"     in wb.sheetnames

    def test_comparables_sheet_has_data(self, sample_df):
        import openpyxl
        result = to_excel(sample_df, self._make_stats(sample_df))
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Comparables"]
        # Row 1 = title, row 3 = header, rows 4+ = data
        # Check that at least one cell in row 4 has "Sunrise Villa"
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "Sunrise Villa" in values

    def test_summary_sheet_has_kpis(self, sample_df):
        import openpyxl
        result = to_excel(sample_df, self._make_stats(sample_df))
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        values = [cell.value for row in ws.iter_rows() for cell in row
                  if cell.value is not None]
        assert "Total Records"       in values
        assert "Average Sale Price"  in values
        assert "Records by Parish"   in values

    def test_empty_dataframe_no_crash(self):
        df_empty = pd.DataFrame(columns=["price_sold", "sq_ft", "parish", "type"])
        stats = summary_stats(df_empty)
        result = to_excel(df_empty, stats)
        assert isinstance(result, bytes)
        assert len(result) > 0


# ═══════════════════════════════════════════════════════════════════════════════
# query() – filter logic against the test database
# ═══════════════════════════════════════════════════════════════════════════════

class TestRunQuery:
    """Tests for run_query() using the shared test database."""

    def test_no_filters_returns_all(self, mem_db):
        df = run_query(mem_db, {})
        assert len(df) == len(SAMPLE_ROWS)

    def test_filter_by_parish(self, mem_db):
        df = run_query(mem_db, {"parishes": ["Pembroke"]})
        assert len(df) == 1
        assert df.iloc[0]["parish"] == "Pembroke"

    def test_filter_multiple_parishes(self, mem_db):
        df = run_query(mem_db, {"parishes": ["Paget", "Warwick"]})
        assert len(df) == 2
        assert set(df["parish"]) == {"Paget", "Warwick"}

    def test_filter_by_type(self, mem_db):
        df = run_query(mem_db, {"types": ["House"]})
        assert all(row == "House" for row in df["type"])
        assert len(df) == 3   # Sunrise Villa, Sea Breeze, No Price Prop

    def test_filter_by_price_min(self, mem_db):
        df = run_query(mem_db, {"price_min": 1_000_000})
        assert all(row >= 1_000_000 for row in df["price_sold"].dropna())
        # Sea Breeze (1.2M) and Commercial (1.5M)
        assert len(df) == 2

    def test_filter_by_price_max(self, mem_db):
        df = run_query(mem_db, {"price_max": 500_000})
        assert all(row <= 500_000 for row in df["price_sold"].dropna())

    def test_filter_price_range(self, mem_db):
        df = run_query(mem_db, {"price_min": 600_000, "price_max": 1_000_000})
        prices = df["price_sold"].dropna().tolist()
        assert all(600_000 <= p <= 1_000_000 for p in prices)

    def test_filter_by_date_from(self, mem_db):
        df = run_query(mem_db, {"date_from": date(2024, 1, 1)})
        for d in df["sale_date"].dropna():
            assert d >= "2024-01-01"

    def test_filter_by_date_to(self, mem_db):
        df = run_query(mem_db, {"date_to": date(2023, 12, 31)})
        for d in df["sale_date"].dropna():
            assert d <= "2023-12-31"

    def test_filter_date_range(self, mem_db):
        df = run_query(mem_db, {"date_from": date(2024, 1, 1), "date_to": date(2024, 12, 31)})
        for d in df["sale_date"].dropna():
            assert "2024-01-01" <= d <= "2024-12-31"
        # Sunrise Villa (2024-03-15) and Harbour View (2024-07-20)
        assert len(df) == 2

    def test_filter_by_sqft_min(self, mem_db):
        df = run_query(mem_db, {"sqft_min": 2000})
        sqfts = df["sq_ft"].dropna()
        assert all(s >= 2000 for s in sqfts)

    def test_filter_beds_min(self, mem_db):
        df = run_query(mem_db, {"beds_min": 4})
        beds = df["beds"].dropna()
        assert all(b >= 4 for b in beds)

    def test_filter_pool_yes(self, mem_db):
        df = run_query(mem_db, {"pool": "Y"})
        assert all(row == "Y" for row in df["pool"])
        # Sunrise Villa and Sea Breeze have pool=Y
        assert len(df) == 2

    def test_filter_pool_no(self, mem_db):
        df = run_query(mem_db, {"pool": "N"})
        assert all(row == "N" for row in df["pool"])

    def test_filter_guest_yes(self, mem_db):
        df = run_query(mem_db, {"guest": "Y"})
        assert len(df) == 1
        assert df.iloc[0]["property_name"] == "Harbour View"

    def test_filter_waterfront_yes(self, mem_db):
        df = run_query(mem_db, {"waterfront": "Y"})
        names = set(df["property_name"])
        assert "Harbour View" in names
        assert "Sea Breeze"   in names

    def test_filter_all_pool_is_noop(self, mem_db):
        df_all  = run_query(mem_db, {})
        df_pool = run_query(mem_db, {"pool": "All"})
        assert len(df_all) == len(df_pool)

    def test_free_text_search_name(self, mem_db):
        df = run_query(mem_db, {"search": "Sunrise"})
        assert len(df) == 1
        assert df.iloc[0]["property_name"] == "Sunrise Villa"

    def test_free_text_search_address(self, mem_db):
        df = run_query(mem_db, {"search": "Harbour Lane"})
        assert len(df) == 1
        assert df.iloc[0]["property_name"] == "Harbour View"

    def test_free_text_search_case_insensitive(self, mem_db):
        df_lower = run_query(mem_db, {"search": "sunrise"})
        df_upper = run_query(mem_db, {"search": "SUNRISE"})
        assert len(df_lower) == len(df_upper) == 1

    def test_no_match_returns_empty(self, mem_db):
        df = run_query(mem_db, {"search": "XXXXXXXXXNOTFOUND"})
        assert len(df) == 0
        assert isinstance(df, pd.DataFrame)

    def test_combined_filters(self, mem_db):
        # House type + pool=Y  →  Sunrise Villa AND Sea Breeze (both House + pool=Y)
        df = run_query(mem_db, {"types": ["House"], "pool": "Y"})
        assert len(df) == 2
        names = set(df["property_name"])
        assert "Sunrise Villa" in names
        assert "Sea Breeze" in names

    def test_results_ordered_by_date_desc(self, mem_db):
        df = run_query(mem_db, {})
        dates = df["sale_date"].dropna().tolist()
        assert dates == sorted(dates, reverse=True)

    def test_result_columns(self, mem_db):
        df = run_query(mem_db, {})
        expected_cols = {
            "property_name", "address", "parish", "type",
            "price_sold", "sale_date", "sq_ft",
            "pool", "guest", "waterfront", "listed",
            "zone", "country",
        }
        assert expected_cols.issubset(set(df.columns))

    def test_filter_units_min(self, mem_db):
        df = run_query(mem_db, {"units_min": 2})
        units = df["units"].dropna()
        assert all(u >= 2 for u in units)
        # Sea Breeze (2 units), Commercial (5 units)
        assert len(df) == 2

    def test_filter_zone(self, mem_db):
        df = run_query(mem_db, {"zones": ["120"]})
        assert len(df) == 1
        assert df.iloc[0]["property_name"] == "Harbour View"

    def test_country_filter(self, mem_db):
        df = run_query(mem_db, {"country": "Bermuda"})
        assert len(df) == len(SAMPLE_ROWS)

    def test_country_filter_no_match(self, mem_db):
        df = run_query(mem_db, {"country": "NonExistentCountry"})
        assert len(df) == 0
